"""
excel_to_oscal.py v2

Converts a legacy Excel System Security Plan (SSP) to OSCAL JSON format.

Key features over a naive conversion:
  - by-components structure: each tool gets its own evidence slot per control
  - UUID v5 deterministic identifiers: stable across runs, clean Git diffs
  - Dynamic tool-to-control mapping: findings map at ingest time, not pre-declared
  - Framework crosswalk support: map high-watermark controls between frameworks
  - GRC agent review flags: surface missing/stale statements automatically

Usage:
    python scripts/excel_to_oscal.py \\
        --input path/to/your-ssp.xlsx \\
        --output oscal/sspp.json \\
        --crosswalk crosswalk/framework-crosswalk.json

Requirements:
    pip install openpyxl
"""

import json
import argparse
import uuid
import re
import sys
sys.stdout.reconfigure(encoding='utf-8')
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.")
    print("Run: pip install openpyxl")
    exit(1)


# ── UUID v5 Configuration ─────────────────────────────────────────────────────
# Deterministic UUIDs: same control always gets same UUID.
# This keeps Git diffs showing only content changes, not ID churn.
#
# The namespace value below is a fixed, non-secret constant used to derive
# deterministic UUIDs via the UUID v5 algorithm. It does not need to be unique
# per organization — any stable, non-rotating UUID works as a namespace.
# The value here is one of the well-known UUIDs from RFC 4122 (DNS namespace).
# You may use any fixed UUID you choose; what matters is that it never changes
# across runs so that the same control always produces the same output UUID.

OSCAL_NAMESPACE = uuid.UUID("6ba7b810-9dad-11d1-80b4-00c04fd430c8")

def stable_uuid(name: str) -> str:
    """
    Generate a deterministic UUID v5 from a name string.
    Same name always produces same UUID — critical for clean Git diffs.

    Examples:
        stable_uuid("control:ac-2")        -> always same UUID
        stable_uuid("component:sailpoint")  -> always same UUID
        stable_uuid("req:ac-2:sailpoint")   -> always same UUID
    """
    return str(uuid.uuid5(OSCAL_NAMESPACE, name))


# ── Tool-to-Control Mapping ───────────────────────────────────────────────────
# Defines which tools provide evidence for which specific control IDs.
# NOTE: Use specific control IDs, NOT control families.
# Family-level mappings cause false positives (see docs/ARCHITECTURE.md).
#
# Add or modify tool entries to match your organization's tool stack.
# Each tool entry requires:
#   - title: display name
#   - type: "software" | "service" | "hardware"
#   - families: [] (leave empty — use specific controls instead)
#   - controls: list of specific NIST 800-53 Rev 5 control IDs this tool evidences
#   - evidence_type: short description of what kind of evidence this produces
#   - api_ready: whether an API connector exists yet
#   - what_it_proves: human-readable description for the OSCAL component catalog

TOOL_CONTROL_COVERAGE = {
    "sailpoint": {
        "title": "SailPoint Identity Security Cloud",
        "type": "software",
        "families": [],
        "controls": [
            "ac-2", "ac-2(2)", "ac-2(3)", "ac-2(4)", "ac-2(5)",
            "ac-2(7)", "ac-2(9)", "ac-2(12)", "ac-2(13)",
            "ac-5", "ac-6", "ac-6(7)",
            "ia-4", "ia-4(4)", "ia-12", "ia-12(1)", "ia-12(2)", "ia-12(3)",
            "ps-4", "ps-5",
        ],
        "evidence_type": "Identity Governance",
        "api_ready": False,
        "what_it_proves": (
            "Account lifecycle (provision/deprovision), access certifications, "
            "SoD violations, inactive account reports, joiner/mover/leaver workflows"
        )
    },
    "forgerock": {
        "title": "ForgeRock / PingOne Advanced Identity Cloud",
        "type": "software",
        "families": [],
        "controls": ["ia-2", "ia-2(1)", "ia-2(2)", "ia-5", "ia-8", "ac-12"],
        "evidence_type": "Authentication & MFA",
        "api_ready": False,
        "what_it_proves": (
            "MFA enrollment rates, AAL2 compliance, session timeout enforcement, "
            "authentication events, failed login attempts"
        )
    },
    "aws_iam": {
        "title": "AWS Identity and Access Management (IAM)",
        "type": "service",
        "families": [],
        "controls": ["ac-2", "ac-3", "ac-6", "ia-2", "ia-5"],
        "evidence_type": "Cloud Identity",
        "api_ready": False,
        "what_it_proves": (
            "IAM role assignments, permission boundaries, MFA on root/privileged accounts, "
            "unused credentials, access key rotation"
        )
    },
    "beyondtrust": {
        "title": "BeyondTrust Privileged Access Management",
        "type": "software",
        "families": [],
        "controls": ["ac-2", "ac-6", "ac-17", "au-2", "ia-5"],
        "evidence_type": "Privileged Access",
        "api_ready": False,
        "what_it_proves": (
            "Privileged session recordings, credential vault status, "
            "JIT access grants, privileged account rotation"
        )
    },
    "wiz": {
        "title": "Wiz Cloud Security Platform",
        "type": "software",
        "families": [],
        "controls": ["ra-5", "cm-6", "cm-7", "sa-11", "si-2", "si-3", "si-4", "sc-7"],
        "evidence_type": "Technical Configuration",
        "api_ready": True,
        "what_it_proves": (
            "Cloud misconfigurations, vulnerability findings, CSPM policy violations, "
            "IaC scan results, container security"
        )
    },
    "tanium": {
        "title": "Tanium Endpoint Management",
        "type": "software",
        "families": [],
        "controls": ["si-2", "cm-8", "ra-5", "cm-3", "sc-28"],
        "evidence_type": "Endpoint Compliance",
        "api_ready": True,
        "what_it_proves": (
            "Patch compliance by severity SLA, software inventory, "
            "endpoint configuration compliance, disk encryption status"
        )
    },
    "jira": {
        "title": "Jira",
        "type": "software",
        "families": [],
        "controls": ["ac-2", "cm-3", "cm-4", "ir-5", "ca-5", "ir-3"],
        "evidence_type": "Process Controls & POA&M",
        "api_ready": True,
        "what_it_proves": (
            "Account request/approval workflows, change management tickets, "
            "POA&M tracking, incident tracking, tabletop exercise records"
        )
    },
    "bitbucket": {
        "title": "Bitbucket",
        "type": "software",
        "families": [],
        "controls": ["cm-3", "cm-5", "sa-10", "au-2"],
        "evidence_type": "Change Control Audit Trail",
        "api_ready": False,
        "what_it_proves": (
            "PR approvals as change control evidence, who approved what and when, "
            "branch protection enforcement, code review completion"
        )
    },
    "terraform": {
        "title": "Terraform",
        "type": "software",
        "families": [],
        "controls": ["cm-2", "cm-3", "cm-6", "sa-10"],
        "evidence_type": "Infrastructure as Code Baseline",
        "api_ready": False,
        "what_it_proves": (
            "Approved infrastructure baseline, change control via PR approval, "
            "configuration drift prevention, version-controlled infra state"
        )
    },
    "ansible": {
        "title": "Ansible",
        "type": "software",
        "families": [],
        "controls": ["cm-2", "cm-6", "cm-7", "si-2"],
        "evidence_type": "Configuration Enforcement",
        "api_ready": False,
        "what_it_proves": (
            "Configuration baseline enforcement, CIS/STIG compliance, "
            "automated remediation runs, patch deployment"
        )
    },
    "splunk": {
        "title": "Splunk SIEM",
        "type": "software",
        "families": [],
        "controls": ["au-2", "au-3", "au-6", "au-9", "au-11", "si-4", "ir-5"],
        "evidence_type": "Audit Logs & SIEM",
        "api_ready": False,
        "what_it_proves": (
            "Audit event collection, log retention compliance, correlation rule coverage, "
            "sensitive data access logs, alert response times"
        )
    },
    "new_relic": {
        "title": "New Relic",
        "type": "software",
        "families": [],
        "controls": ["si-4", "au-6", "ca-7"],
        "evidence_type": "Observability & Anomaly Detection",
        "api_ready": False,
        "what_it_proves": (
            "Application anomaly detection, performance baselines, "
            "availability monitoring, incident detection"
        )
    },
    "aws_storage": {
        "title": "AWS S3 / Aurora / Redshift",
        "type": "service",
        "families": [],
        "controls": ["sc-28", "sc-13", "cp-9", "sc-8"],
        "evidence_type": "Data Encryption & Backup",
        "api_ready": False,
        "what_it_proves": (
            "Encryption at rest (KMS), backup completion, "
            "retention policy compliance, TLS enforcement"
        )
    },
    "goanywhere": {
        "title": "GoAnywhere MFT",
        "type": "software",
        "families": [],
        "controls": ["sc-8", "sc-12", "au-2", "ac-17"],
        "evidence_type": "Secure File Transfer",
        "api_ready": False,
        "what_it_proves": (
            "FIPS 140 encrypted transfers, transfer audit logs, "
            "partner authentication, failed transfer alerts"
        )
    },
    "nexus": {
        "title": "Nexus Sonatype",
        "type": "software",
        "families": [],
        "controls": ["sr-3", "sa-11", "sa-15"],
        "evidence_type": "Software Supply Chain",
        "api_ready": False,
        "what_it_proves": (
            "SBOM completeness, vulnerable dependency blocking, "
            "approved component inventory, license compliance"
        )
    },
    "jenkins": {
        "title": "Jenkins",
        "type": "software",
        "families": [],
        "controls": ["sa-10", "sa-11", "cm-3", "si-2"],
        "evidence_type": "CI/CD Pipeline Security",
        "api_ready": False,
        "what_it_proves": (
            "Security gate enforcement in pipelines, automated test results, "
            "deployment approvals, patch deployment runs"
        )
    },
    "zscaler": {
        "title": "Zscaler",
        "type": "service",
        "families": [],
        "controls": ["sc-7", "ac-17", "si-3", "sc-8"],
        "evidence_type": "Network Security",
        "api_ready": False,
        "what_it_proves": (
            "Boundary protection, zero trust network access, "
            "inline inspection, TLS enforcement"
        )
    },
    "aws_eks": {
        "title": "AWS EKS / Docker",
        "type": "service",
        "families": [],
        "controls": ["cm-7", "sc-39", "cm-8"],
        "evidence_type": "Container Security",
        "api_ready": False,
        "what_it_proves": (
            "Container network policies, process isolation, "
            "image inventory and scanning"
        )
    },
    "blazemeter": {
        "title": "BlazeMeter",
        "type": "software",
        "families": [],
        "controls": ["ca-8", "sa-11"],
        "evidence_type": "Security Testing",
        "api_ready": False,
        "what_it_proves": (
            "Performance and security test results, regression test coverage"
        )
    },
}

# ── Framework crosswalk configuration ────────────────────────────────────────
# Controls that have higher requirements under a secondary framework.
# Used to tag controls in the OSCAL output for dual-framework compliance.
# Populate this list based on your specific framework requirements.
HIGH_WATERMARK_CONTROLS = [
    # Example: controls where secondary framework is stricter than primary
    "ac-2", "ac-6", "ac-17", "ia-2", "ia-5",
    "au-2", "au-9", "ir-3", "ir-6", "sc-7",
    "sc-8", "sc-28", "sr-3",
]

# ── Sheet configuration ───────────────────────────────────────────────────────
# Update these to match your Excel SSP structure.
# SHEET_NAME: the tab name containing control data
# HEADER_ROW: the row number containing column headers
# DATA_START_ROW: the first row containing actual control data

SHEET_NAME = "Mandatory Baseline"  # Update to match your sheet name
HEADER_ROW = 31                    # Update to match your header row
DATA_START_ROW = 32                # Update to match your first data row

# Column positions (1-indexed) — update to match your SSP column layout
COLUMNS = {
    "number":         1,
    "family":         2,
    "control_id":     3,
    "control_name":   4,
    "control_text":   5,
    "supplemental":   6,
    "related":        7,
    "status":         8,
    "implementation": 9,
    "owner2":         10,
    "owner3":         11,
    "owner4":         12,
}

VALID_STATUSES = {
    "implemented":    "implemented",
    "inherited":      "inherited",
    "planned":        "planned",
    "not applicable": "not-applicable",
    "not-applicable": "not-applicable",
    "n/a":            "not-applicable",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def normalize_control_id(raw_id: str) -> str:
    """Normalize control ID to lowercase OSCAL format. AC-02(01) -> ac-2(1)"""
    if not raw_id:
        return None
    s = str(raw_id).strip().lower()
    s = re.sub(r'-0*(\d)', r'-\1', s)
    s = re.sub(r'\(0*(\d+)\)', r'(\1)', s)
    return s


def normalize_status(raw_status: str) -> str:
    if not raw_status:
        return "not-implemented"
    key = str(raw_status).strip().lower()
    return VALID_STATUSES.get(key, "not-implemented")


def get_cell(sheet, row: int, col: int):
    val = sheet.cell(row=row, column=col).value
    if val is None:
        return None
    cleaned = str(val).strip()
    return cleaned if cleaned else None


def get_tools_for_control(control_id: str) -> list:
    """Return list of tool keys that cover this specific control."""
    return [
        tool_key
        for tool_key, tool_data in TOOL_CONTROL_COVERAGE.items()
        if control_id in tool_data["controls"]
    ]


def build_by_components(control_id: str, ssp_text: str, owners: list, status: str) -> list:
    """
    Build the by-components array for a control.

    The SSP narrative goes into the primary system component.
    Each tool that covers this control gets its own component entry.
    API connectors populate tool-specific entries later.

    This structure is the key that makes the reconciler work:
    - Primary component = the claim (SSP narrative)
    - Tool components = the evidence (API data)
    - Reconciler = compares claim vs. evidence
    """
    components = []

    # Primary system component — holds the SSP narrative (the claim)
    primary_desc = ssp_text if ssp_text else (
        "Implementation statement pending. "
        "Will be updated by GRC agent using API evidence from connected tools."
    )

    if len(owners) > 1:
        primary_desc += f"\n\nShared responsibility documented across: {', '.join(owners)}."

    components.append({
        "component-uuid": stable_uuid("component:this-system"),
        "description": primary_desc,
        "implementation-status": {
            "state": status
        },
        "remarks": (
            "Primary implementation narrative from SSP baseline. "
            "Evidence confirmation process: "
            "(1) API connectors run and write findings to tool-specific component slots. "
            "(2) reconcile_oscal.py reads all populated slots, updates this narrative, "
            "and flips implementation-status when evidence supports it. "
            "(3) Human ISSO review required before submission. "
            "Last reconciled: never — run scripts/reconcile_oscal.py to update."
        ),
        "props": [
            {"name": "last-reconciled", "value": "never"},
            {"name": "reconciled-by", "value": "pending"},
            {"name": "evidence-sources-active", "value": "none"}
        ]
    })

    # Tool-specific components — placeholders that API connectors will populate
    covering_tools = get_tools_for_control(control_id)

    for tool_key in covering_tools:
        tool = TOOL_CONTROL_COVERAGE[tool_key]
        api_status = "api-connected" if tool["api_ready"] else "pending-integration"

        if not tool["api_ready"]:
            description = (
                f"{tool['title']} provides {tool['evidence_type']} evidence for this control. "
                f"Specifically: {tool['what_it_proves']}. "
                f"[API EVIDENCE PENDING — connector not yet integrated]"
            )
            impl_state = "planned"
        else:
            description = (
                f"{tool['title']} provides {tool['evidence_type']} evidence for this control. "
                f"Specifically: {tool['what_it_proves']}. "
                f"[API READY — evidence will auto-populate on next connector run]"
            )
            impl_state = "implemented"

        components.append({
            "component-uuid": stable_uuid(f"component:{tool_key}"),
            "description": description,
            "implementation-status": {
                "state": impl_state,
                "remarks": api_status
            },
            "props": [
                {"name": "tool-name", "value": tool_key},
                {"name": "evidence-type", "value": tool["evidence_type"]},
                {"name": "api-ready", "value": str(tool["api_ready"]).lower()},
                {"name": "last-api-pull", "value": "never"}  # Gate: reconciler checks this
            ]
        })

    return components


def is_missing_or_stale(text: str, status: str) -> tuple:
    """Check if an implementation statement needs GRC agent review."""
    if status == "not-applicable":
        return False, None
    if not text:
        return True, "missing"
    if len(text) < 50:
        return True, "too_short"
    placeholders = ["tbd", "to be determined", "placeholder", "n/a", "none", "pending"]
    if text.lower().strip() in placeholders:
        return True, "placeholder"
    return False, None


def load_crosswalk(crosswalk_path: str) -> dict:
    """Load framework crosswalk for high-watermark control tagging."""
    try:
        with open(crosswalk_path, "r") as f:
            data = json.load(f)
        return {item["control-id"]: item for item in data.get("control-deltas", [])}
    except Exception as e:
        print(f"  WARNING: Could not load crosswalk: {e}")
        return {}


# ── Component catalog builder ─────────────────────────────────────────────────

def build_component_catalog() -> list:
    """
    Build the system-implementation components list from the tool stack.
    Every tool in TOOL_CONTROL_COVERAGE becomes a named component.
    This is the authoritative list that by-components entries reference.
    """
    components = [
        {
            "uuid": stable_uuid("component:this-system"),
            "type": "this-system",
            "title": "Primary System",
            "description": "The primary system being assessed.",
            "status": {"state": "operational"}
        }
    ]

    for tool_key, tool in TOOL_CONTROL_COVERAGE.items():
        components.append({
            "uuid": stable_uuid(f"component:{tool_key}"),
            "type": tool["type"],
            "title": tool["title"],
            "description": f"{tool['evidence_type']} — {tool['what_it_proves']}",
            "props": [
                {"name": "tool-key", "value": tool_key},
                {"name": "evidence-type", "value": tool["evidence_type"]},
                {"name": "api-ready", "value": str(tool["api_ready"]).lower()},
            ],
            "status": {"state": "operational"}
        })

    return components


# ── Main converter ────────────────────────────────────────────────────────────

def convert_excel_to_oscal(input_path: str, output_path: str, crosswalk_path: str = None):
    print(f"\n{'='*60}")
    print(f"  Excel SSP -> OSCAL Converter v2")
    print(f"  by-components structure + UUID v5 identifiers")
    print(f"{'='*60}")
    print(f"  Input:     {input_path}")
    print(f"  Output:    {output_path}")
    print(f"{'='*60}\n")

    crosswalk = {}
    if crosswalk_path and Path(crosswalk_path).exists():
        crosswalk = load_crosswalk(crosswalk_path)
        print(f"  Loaded crosswalk: {len(crosswalk)} control deltas\n")

    print(f"  Opening Excel file...")
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"  ERROR: Could not open Excel file: {e}")
        return False

    if SHEET_NAME not in wb.sheetnames:
        print(f"  ERROR: Sheet '{SHEET_NAME}' not found.")
        print(f"  Available sheets: {wb.sheetnames}")
        return False

    ws = wb[SHEET_NAME]
    print(f"  Found sheet: {SHEET_NAME}")
    print(f"  Processing controls from row {DATA_START_ROW}...\n")

    implemented_requirements = []
    stats = {
        "total": 0,
        "implemented": 0,
        "inherited": 0,
        "planned": 0,
        "not_applicable": 0,
        "missing_statements": 0,
        "stale_statements": 0,
        "high_watermark": 0,
        "multi_tool_controls": 0,
    }

    for row_num in range(DATA_START_ROW, ws.max_row + 1):
        row_data = {key: get_cell(ws, row_num, col) for key, col in COLUMNS.items()}

        if not row_data["control_id"]:
            continue

        stats["total"] += 1
        control_id = normalize_control_id(row_data["control_id"])
        if not control_id:
            continue

        status = normalize_status(row_data["status"])

        # Collect all owner narratives (shared responsibility columns)
        owners = []
        texts = []
        if row_data["implementation"]:
            owners.append("Primary Owner")
            texts.append(row_data["implementation"])
        for i, key in enumerate(["owner2", "owner3", "owner4"], 2):
            if row_data[key]:
                owners.append(f"Owner {i}")
                texts.append(f"[Owner {i}]: {row_data[key]}")
        ssp_text = "\n\n".join(texts) if texts else None

        needs_review, review_reason = is_missing_or_stale(ssp_text, status)
        is_high_watermark = control_id in HIGH_WATERMARK_CONTROLS
        crosswalk_data = crosswalk.get(control_id, {})
        covering_tools = get_tools_for_control(control_id)

        # Update stats
        stat_map = {
            "implemented": "implemented",
            "inherited": "inherited",
            "planned": "planned",
            "not-applicable": "not_applicable"
        }
        if status in stat_map:
            stats[stat_map[status]] += 1
        if needs_review and review_reason == "missing":
            stats["missing_statements"] += 1
        elif needs_review:
            stats["stale_statements"] += 1
        if is_high_watermark:
            stats["high_watermark"] += 1
        if len(covering_tools) > 1:
            stats["multi_tool_controls"] += 1

        # Build OSCAL props for this control
        family = control_id.split("-")[0] if "-" in control_id else ""
        props = [
            {"name": "control-origination", "value": status},
            {"name": "control-family", "value": family.upper()},
        ]

        if is_high_watermark and crosswalk_data:
            props.append({
                "name": "high-watermark",
                "value": crosswalk_data.get("watermark", "secondary-framework"),
                "remarks": crosswalk_data.get("ssp-note", "")
            })
            props.append({
                "name": "gap-risk",
                "value": crosswalk_data.get("gap-risk", "medium")
            })

        if needs_review:
            props.append({
                "name": "grc-agent-review-needed",
                "value": "true",
                "remarks": f"Reason: {review_reason}. Agent will draft using API evidence."
            })

        if covering_tools:
            props.append({
                "name": "evidence-sources",
                "value": ", ".join(covering_tools)
            })

        # Build the implemented-requirement with by-components structure
        req = {
            "uuid": stable_uuid(f"req:{control_id}"),
            "control-id": control_id,
            "props": props,
            "statements": [
                {
                    "statement-id": f"{control_id}_stmt",
                    "uuid": stable_uuid(f"stmt:{control_id}"),
                    "by-components": build_by_components(
                        control_id, ssp_text, owners, status
                    )
                }
            ]
        }

        if row_data["supplemental"]:
            req["remarks"] = f"Supplemental Guidance: {row_data['supplemental']}"

        if row_data["related"]:
            req["related-controls"] = row_data["related"]

        implemented_requirements.append(req)

        if stats["total"] % 50 == 0:
            print(f"  Processed {stats['total']} controls...")

    print(f"\n  Processed {stats['total']} controls total.\n")

    # Build full OSCAL SSP document
    now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

    oscal_ssp = {
        "system-security-plan": {
            "uuid": stable_uuid("document:system-sspp"),
            "metadata": {
                "title": "System Security and Privacy Plan (SSPP)",
                "last-modified": now,
                "version": "2.0",
                "oscal-version": "1.1.2",
                "remarks": (
                    "Generated by excel_to_oscal.py v2. "
                    "Uses by-components structure for shared responsibility. "
                    "Uses UUID v5 for stable identifiers across runs. "
                    "Controls with grc-agent-review-needed=true require implementation "
                    "statement updates from API evidence."
                )
            },
            "import-profile": {
                "href": "https://csrc.nist.gov/projects/risk-management/sp800-53-controls/release-search#/controls?version=5.1",
                "remarks": "NIST SP 800-53 Rev 5 control catalog"
            },
            "system-characteristics": {
                "system-name": "System Name",
                "system-name-short": "SYS",
                "description": "System description — update with your system's purpose and scope.",
                "security-sensitivity-level": "moderate",
                "system-information": {
                    "information-types": [
                        {
                            "uuid": stable_uuid("infotype:primary"),
                            "title": "Primary Information Type",
                            "description": "Update with your system's primary information type.",
                            "confidentiality-impact": {"base": "moderate"},
                            "integrity-impact": {"base": "moderate"},
                            "availability-impact": {"base": "moderate"}
                        }
                    ]
                },
                "security-impact-level": {
                    "security-objective-confidentiality": "moderate",
                    "security-objective-integrity": "moderate",
                    "security-objective-availability": "moderate"
                },
                "status": {"state": "operational"},
                "authorization-boundary": {
                    "description": "Update with your system's authorization boundary description."
                }
            },
            "system-implementation": {
                "remarks": (
                    "Each component corresponds to a tool that provides evidence for one or more controls. "
                    "API connectors populate evidence into each component's by-components entry."
                ),
                "users": [
                    {
                        "uuid": stable_uuid("user:isso"),
                        "title": "Information System Security Officer (ISSO)",
                        "role-ids": ["isso"]
                    },
                    {
                        "uuid": stable_uuid("user:system-owner"),
                        "title": "System Owner",
                        "role-ids": ["system-owner"]
                    },
                    {
                        "uuid": stable_uuid("user:admin"),
                        "title": "System Administrator",
                        "role-ids": ["admin"]
                    }
                ],
                "components": build_component_catalog()
            },
            "control-implementation": {
                "description": (
                    "Implementation of NIST SP 800-53 Rev 5 mandatory baseline. "
                    "Each control uses by-components structure to capture shared responsibility. "
                    "API connectors update tool-specific component entries with live evidence."
                ),
                "implemented-requirements": implemented_requirements
            }
        }
    }

    # Write output
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(oscal_ssp, f, indent=2, ensure_ascii=False)

    # Summary report
    coverage_pct = round(
        (stats["implemented"] + stats["inherited"]) / stats["total"] * 100, 1
    ) if stats["total"] > 0 else 0

    print(f"{'='*60}")
    print(f"  CONVERSION COMPLETE")
    print(f"{'='*60}")
    print(f"  Output:                {output_path}")
    print(f"  UUID strategy:         v5 (deterministic — stable across runs)")
    print(f"  Structure:             by-components (shared responsibility)")
    print(f"{'─'*60}")
    print(f"  Total controls:        {stats['total']}")
    print(f"  Implemented:           {stats['implemented']}")
    print(f"  Inherited:             {stats['inherited']}")
    print(f"  Planned:               {stats['planned']}")
    print(f"  Not Applicable:        {stats['not_applicable']}")
    print(f"  Coverage:              {coverage_pct}%")
    print(f"{'─'*60}")
    print(f"  Multi-tool controls:   {stats['multi_tool_controls']}")
    print(f"  High-watermark tagged: {stats['high_watermark']}")
    print(f"{'─'*60}")
    print(f"  GRC Agent Review Needed:")
    print(f"    Missing statements:  {stats['missing_statements']}")
    print(f"    Stale statements:    {stats['stale_statements']}")
    print(f"    Total for review:    {stats['missing_statements'] + stats['stale_statements']}")
    print(f"{'='*60}")
    print(f"\n  Next step: Run API connectors to populate by-components evidence slots.")
    print(f"  Tools API-ready: {[k for k, v in TOOL_CONTROL_COVERAGE.items() if v['api_ready']]}\n")

    return True


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert Excel SSP to OSCAL JSON (v2 — by-components + UUID v5)"
    )
    parser.add_argument("--input", required=True, help="Path to Excel SSP (.xlsx)")
    parser.add_argument("--output", default="oscal/sspp.json", help="Output OSCAL JSON path")
    parser.add_argument(
        "--crosswalk",
        default="crosswalk/framework-crosswalk.json",
        help="Framework crosswalk JSON path"
    )

    args = parser.parse_args()
    convert_excel_to_oscal(args.input, args.output, args.crosswalk)
